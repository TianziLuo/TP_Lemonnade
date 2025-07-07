import os
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import win32com.client as win32

def SKU_out():
    # ========= Path configuration =========
    base_name      = "上传易仓SKU映射关系"
    downloads_dir  = os.path.join(os.path.expanduser("~"), "Downloads")
    xlsx_out_path  = os.path.join(downloads_dir, f"{base_name}.xlsx")
    xls_out_path   = os.path.join(downloads_dir, f"{base_name}.xls")

    SOURCE_PATH    = r"C:\Frank\1.1_核心.xlsx"
    TEMPLATE_PATH  = r"C:\Template\上传易仓SKU映射关系.xlsx"
    SHEET_NAME     = "上传易仓SKU映射关系"

    # ========= Safety: ensure the source file was saved within the last 30 seconds =========
    MAX_FILE_AGE  = time.time() - os.path.getmtime(SOURCE_PATH)
    if MAX_FILE_AGE > 30:
        raise RuntimeError(
            f"⚠️ Source file '{SOURCE_PATH}' was last saved {int(MAX_FILE_AGE)} seconds ago, exceeding the 30-second limit. Please save again and retry."
        )

    # ========= Load workbooks =========
    src_wb  = load_workbook(SOURCE_PATH, data_only=True)
    src_ws  = src_wb[SHEET_NAME]
    tmpl_wb = load_workbook(TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    # ========= Define copy range =========
    START_COL  = 1   # Column A
    END_COL    = 6   # Column F
    START_ROW  = 2
    PASTE_ROW0 = 2

    # Determine the last non‑empty row in the source sheet
    last_row = src_ws.max_row
    while last_row >= START_ROW and all(
        src_ws.cell(row=last_row, column=col).value is None
        for col in range(START_COL, END_COL + 1)
    ):
        last_row -= 1

    # ========= Define orange fill style =========
    orange_fill = PatternFill(
        fill_type="solid",
        start_color="FFFDE9D9",
        end_color="FFFDE9D9",
    )

    # ========= Copy values and apply fill =========
    for row_src, row_dst in zip(
        range(START_ROW, last_row + 1),
        range(PASTE_ROW0, PASTE_ROW0 + (last_row - START_ROW + 1)),
    ):
        for col in range(START_COL, END_COL + 1):
            src_val  = src_ws.cell(row=row_src, column=col).value
            tgt_cell = tmpl_ws.cell(row=row_dst, column=col)

            # Keep empty cells as empty strings; strip whitespace otherwise
            tgt_cell.value = "" if src_val is None else str(src_val).strip()

            # Apply background color to columns A–C
            if col <= 3:
                tgt_cell.fill = orange_fill

    # ========= Save the .xlsx file =========
    # If an old version of the file exists, remove it first to avoid Excel opening an outdated copy
    if os.path.exists(xlsx_out_path):
        os.remove(xlsx_out_path)

    tmpl_wb.save(xlsx_out_path)
    print("✅ Saved SKU.xlsx")

    # ========= Use Excel COM to save as .xls =========
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible       = False
    excel.DisplayAlerts = False

    try:
        wb_excel = excel.Workbooks.Open(xlsx_out_path)
        wb_excel.SaveAs(xls_out_path, FileFormat=56)  # 56 = .xls
        wb_excel.Close()
    finally:
        excel.Quit()

    print("✅ Saved SKU.xls")

    # ========= Delete the intermediate .xlsx file from Downloads =========
    if os.path.exists(xlsx_out_path):
        os.remove(xlsx_out_path)
        print("🗑️ Deleted SKU.xlsx")

'''
if __name__ == "__main__":
    SKU_out()
'''