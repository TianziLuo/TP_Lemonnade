from openpyxl import load_workbook
import pandas as pd
from pathlib import Path

def generate_tp_csv(source_path, sheet_name, template_path, output_path):
    df = pd.read_excel(source_path, sheet_name=sheet_name, header=None, skiprows=1, usecols="A:F")
    df = df.dropna(how='all')

    if df.shape[1] >= 2:
        df.iloc[:, 1] = df.iloc[:, 1].apply(lambda x: f"'{x}" if pd.notna(x) else x)

    with open(template_path, "r", encoding="utf-8") as f:
        header = f.readline().strip()

    with open(output_path, "w", encoding="utf-8", newline="") as f:
        f.write(header + "\n")
        df.to_csv(f, index=False, header=False)

    print(f"✅ TP CSV 已生成: {output_path}")


def update_inventory_excel(source_path, sheet_name, template_path, output_path,
                            start_col=1, end_col=7, start_row=2, paste_start_row=2):
    source_wb = load_workbook(source_path, data_only=True)
    source_ws = source_wb[sheet_name]

    template_wb = load_workbook(template_path)
    template_ws = template_wb.active

    last_row = source_ws.max_row
    while last_row >= start_row and all(source_ws.cell(row=last_row, column=col).value is None for col in range(start_col, end_col + 1)):
        last_row -= 1

    for idx, row in enumerate(range(start_row, last_row + 1), start=paste_start_row):
        for col in range(start_col, end_col + 1):
            template_ws.cell(row=idx, column=col - start_col + 1).value = source_ws.cell(row=row, column=col).value

    template_wb.save(output_path)
    print(f"✅ 店小秘库存表已保存: {output_path}")

'''
if __name__ == "__main__":
    downloads = Path.home() / "Downloads"

    update_inventory_excel(
    source_path=r"C:\Frank\2.2_店小秘.xlsx",
    sheet_name="盘点",
    template_path=r"C:\Template\店小秘 更新库存.xlsx",
    output_path=downloads / "店小秘 更新库存.xlsx"
)
'''